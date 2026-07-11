import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pyodbc
import pandas as pd
from tkcalendar import DateEntry
import datetime
import os
import sys


from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SERVER = 'SQL IP ADRESİNİZ'
DATABASE = 'MIKRO VERITABANINIZ'
USERNAME = 'SQL Username'
PASSWORD = 'User Password'
# --------------------------------

def kaynak_yolu(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

class StokRaporApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Mikro Stok Hareket Raporu")
        self.root.geometry("1100x650") 
        self.df = None

        # Logo
        try:
            self.root.iconbitmap(kaynak_yolu("logo.ico"))
        except Exception:
            pass

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview", rowheight=25, borderwidth=1, relief="solid")
        style.configure("Treeview.Heading", font=('Helvetica', 9, 'bold'), borderwidth=1, relief="solid")

        # Üst Panel
        top_frame = tk.Frame(root, pady=10, padx=10)
        top_frame.pack(fill=tk.X)

        bugun = datetime.date.today()
        pazartesi = bugun - datetime.timedelta(days=bugun.weekday()) 
        cuma = pazartesi + datetime.timedelta(days=4)                

        tk.Label(top_frame, text="Başlangıç:").pack(side=tk.LEFT, padx=5)
        self.entry_start = DateEntry(top_frame, width=12, background='darkblue', 
                                     foreground='white', borderwidth=2, date_pattern='dd-mm-yyyy', locale='tr_TR')
        self.entry_start.set_date(pazartesi)
        self.entry_start.pack(side=tk.LEFT, padx=5)

        tk.Label(top_frame, text="Bitiş:").pack(side=tk.LEFT, padx=5)
        self.entry_end = DateEntry(top_frame, width=12, background='darkblue', 
                                   foreground='white', borderwidth=2, date_pattern='dd-mm-yyyy', locale='tr_TR')
        self.entry_end.set_date(cuma)
        self.entry_end.pack(side=tk.LEFT, padx=5)

        btn_raporla = tk.Button(top_frame, text="Raporla", command=self.rapor_getir, bg="lightblue")
        btn_raporla.pack(side=tk.LEFT, padx=15)

        btn_excel = tk.Button(top_frame, text="Excel'e Aktar", command=self.excele_aktar, bg="lightgreen")
        btn_excel.pack(side=tk.LEFT, padx=5)

        btn_excel_ozet = tk.Button(top_frame, text="Sadece Genel Toplamı Excel'e Aktar", command=self.excele_aktar_ozet, bg="#ffd700")
        btn_excel_ozet.pack(side=tk.LEFT, padx=5)

        # Alt Panel
        bottom_frame = tk.Frame(root, pady=10, padx=10, bg="#e8e8e8", relief=tk.SUNKEN, borderwidth=2)
        bottom_frame.pack(side=tk.BOTTOM, fill=tk.X)

        self.lbl_donem_basi = tk.Label(bottom_frame, text="DÖNEM BAŞI: 0.00", font=("Arial", 10, "bold"), bg="#e8e8e8")
        self.lbl_donem_basi.pack(side=tk.LEFT, expand=True)

        self.lbl_giris = tk.Label(bottom_frame, text="GİRİŞ: 0.00", font=("Arial", 10, "bold"), bg="#e8e8e8", fg="green")
        self.lbl_giris.pack(side=tk.LEFT, expand=True)

        self.lbl_cikis = tk.Label(bottom_frame, text="ÇIKIŞ: 0.00", font=("Arial", 10, "bold"), bg="#e8e8e8", fg="red")
        self.lbl_cikis.pack(side=tk.LEFT, expand=True)

        self.lbl_net = tk.Label(bottom_frame, text="NET: 0.00", font=("Arial", 10, "bold"), bg="#e8e8e8")
        self.lbl_net.pack(side=tk.LEFT, expand=True)

        self.lbl_donem_sonu = tk.Label(bottom_frame, text="DÖNEM SONU: 0.00", font=("Arial", 10, "bold"), bg="#e8e8e8", fg="blue")
        self.lbl_donem_sonu.pack(side=tk.LEFT, expand=True)

        # Tablo
        tree_frame = tk.Frame(root)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        tree_scroll = tk.Scrollbar(tree_frame)
        tree_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        self.tree = ttk.Treeview(tree_frame, yscrollcommand=tree_scroll.set, show='headings')
        self.tree.pack(fill=tk.BOTH, expand=True)
        tree_scroll.config(command=self.tree.yview)

    def rapor_getir(self):
        start_date_obj = self.entry_start.get_date()
        end_date_obj = self.entry_end.get_date()

        start_date_sql = start_date_obj.strftime('%Y-%m-%d 00:00:00')
        end_date_sql = end_date_obj.strftime('%Y-%m-%d 23:59:59')

        conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={SERVER};DATABASE={DATABASE};UID={USERNAME};PWD={PASSWORD}"
        
        sql = """
        WITH Hareket_Ozet AS (
            SELECT 
                sth_stok_kod,
                SUM(CASE 
                    WHEN sth_tarih < ? THEN 
                        (CASE WHEN (sth_tip = 0 OR sth_tip = 2) AND sth_giris_depo_no = 1 THEN sth_miktar ELSE 0 END) - 
                        (CASE WHEN (sth_tip = 1 OR sth_tip = 2) AND sth_cikis_depo_no = 1 THEN sth_miktar ELSE 0 END)
                    ELSE 0 
                END) AS Donem_Basi,
                
                SUM(CASE 
                    WHEN sth_tarih BETWEEN ? AND ? AND (sth_tip = 0 OR sth_tip = 2) AND sth_giris_depo_no = 1 THEN sth_miktar 
                    ELSE 0 
                END) AS Iceri_Giris,
                
                SUM(CASE 
                    WHEN sth_tarih BETWEEN ? AND ? AND (sth_tip = 1 OR sth_tip = 2) AND sth_cikis_depo_no = 1 THEN sth_miktar 
                    ELSE 0 
                END) AS Disari_Cikis
                
            FROM STOK_HAREKETLERI WITH (NOLOCK)
            WHERE sth_tarih <= ?
              AND sth_iptal = 0
              AND (
                  ((sth_tip = 0 OR sth_tip = 2) AND sth_giris_depo_no = 1) OR 
                  ((sth_tip = 1 OR sth_tip = 2) AND sth_cikis_depo_no = 1)
              )
            GROUP BY sth_stok_kod
        )
        SELECT 
            STK.sto_kod AS [Stok Kodu],
            STK.sto_isim AS [Stok Adı],
            O.Donem_Basi AS [Dönem Başı Stok],
            O.Iceri_Giris AS [Dönem İçi Giriş],
            O.Disari_Cikis AS [Dönem İçi Çıkış],
            (O.Iceri_Giris - O.Disari_Cikis) AS [Net Hareket],
            (O.Donem_Basi + O.Iceri_Giris - O.Disari_Cikis) AS [Dönem Sonu Stok]
        FROM Hareket_Ozet O
        LEFT JOIN STOKLAR STK WITH (NOLOCK) ON O.sth_stok_kod = STK.sto_kod
        WHERE (O.Donem_Basi <> 0 OR O.Iceri_Giris <> 0 OR O.Disari_Cikis <> 0)
        ORDER BY STK.sto_kod;
        """
        
        params = (start_date_sql, start_date_sql, end_date_sql, start_date_sql, end_date_sql, end_date_sql)

        try:
            conn = pyodbc.connect(conn_str)
            self.df = pd.read_sql(sql, conn, params=params)
            conn.close()
            self.tabloyu_guncelle()
            self.toplamlari_hesapla()
        except Exception as e:
            messagebox.showerror("Hata", f"Veritabanı hatası:\n{str(e)}")

    def tabloyu_guncelle(self):
        self.tree.delete(*self.tree.get_children())
        
        if self.df is None or self.df.empty:
            messagebox.showinfo("Bilgi", "Bu tarih aralığında kayıt bulunamadı.")
            return

        self.tree["columns"] = list(self.df.columns)
        
        col_widths = {
            "Stok Kodu": 130,
            "Stok Adı": 350,
            "Dönem Başı Stok": 120,
            "Dönem İçi Giriş": 120,
            "Dönem İçi Çıkış": 120,
            "Net Hareket": 120,
            "Dönem Sonu Stok": 120
        }

        for col in self.tree["columns"]:
            self.tree.heading(col, text=col, anchor=tk.CENTER)
            width = col_widths.get(col, 120)
            self.tree.column(col, width=width, anchor=tk.CENTER)

        for index, row in self.df.iterrows():
            self.tree.insert("", "end", values=list(row))

    def toplamlari_hesapla(self):
        if self.df is not None and not self.df.empty:
            t_donem_basi = self.df['Dönem Başı Stok'].sum()
            t_giris = self.df['Dönem İçi Giriş'].sum()
            t_cikis = self.df['Dönem İçi Çıkış'].sum()
            t_net = self.df['Net Hareket'].sum()
            t_donem_sonu = self.df['Dönem Sonu Stok'].sum()

            self.lbl_donem_basi.config(text=f"DÖNEM BAŞI: {t_donem_basi:,.2f}")
            self.lbl_giris.config(text=f"GİRİŞ: {t_giris:,.2f}")
            self.lbl_cikis.config(text=f"ÇIKIŞ: {t_cikis:,.2f}")
            self.lbl_net.config(text=f"NET: {t_net:,.2f}")
            self.lbl_donem_sonu.config(text=f"DÖNEM SONU: {t_donem_sonu:,.2f}")
        else:
            self.lbl_donem_basi.config(text="DÖNEM BAŞI: 0.00")
            self.lbl_giris.config(text="GİRİŞ: 0.00")
            self.lbl_cikis.config(text="ÇIKIŞ: 0.00")
            self.lbl_net.config(text="NET: 0.00")
            self.lbl_donem_sonu.config(text="DÖNEM SONU: 0.00")

    def excele_aktar(self):
        if self.df is None or self.df.empty:
            messagebox.showwarning("Uyarı", "Dışa aktarılacak veri yok. Önce raporu çalıştırın.")
            return

        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                 filetypes=[("Excel Dosyası", "*.xlsx")],
                                                 title="Excel Olarak Kaydet")
        if file_path:
            try:
                start_dt = self.entry_start.get_date().strftime('%d-%m-%Y')
                end_dt = self.entry_end.get_date().strftime('%d-%m-%Y')
                
                toplamlar = pd.DataFrame([{
                    "Stok Kodu": "GENEL TOPLAM",
                    "Stok Adı": "",
                    "Dönem Başı Stok": self.df['Dönem Başı Stok'].sum(),
                    "Dönem İçi Giriş": self.df['Dönem İçi Giriş'].sum(),
                    "Dönem İçi Çıkış": self.df['Dönem İçi Çıkış'].sum(),
                    "Net Hareket": self.df['Net Hareket'].sum(),
                    "Dönem Sonu Stok": self.df['Dönem Sonu Stok'].sum()
                }])
                df_excel = pd.concat([self.df, toplamlar], ignore_index=True)

                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    df_excel.to_excel(writer, sheet_name='Depo Durum Raporu', index=False, startrow=3)
                    
                    workbook = writer.book
                    worksheet = writer.sheets['Depo Durum Raporu']
                    worksheet.views.sheetView[0].showGridLines = True
                    
                    worksheet['A1'] = "HAFTALIK DEPO DURUM RAPORU"
                    worksheet['A1'].font = Font(name='Segoe UI', size=16, bold=True, color='1F497D')
                    
                    worksheet['A2'] = f"Rapor Tarih Aralığı: {start_dt} - {end_dt}"
                    worksheet['A2'].font = Font(name='Segoe UI', size=11, italic=True, color='595959')
                    
                    header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
                    header_fill = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')
                    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    
                    data_font = Font(name='Segoe UI', size=10)
                    center_align = Alignment(horizontal='center', vertical='center')
                    left_align = Alignment(horizontal='left', vertical='center')
                    right_align = Alignment(horizontal='right', vertical='center')
                    
                    thin_border = Side(border_style="thin", color="D9D9D9")
                    data_border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
                    
                    total_font = Font(name='Segoe UI', size=11, bold=True, color='000000')
                    total_fill = PatternFill(start_color='E9EDF4', end_color='E9EDF4', fill_type='solid')
                    total_border = Border(top=Side(border_style="thin", color="1F497D"), 
                                         bottom=Side(border_style="double", color="1F497D"))
                    
                    worksheet.row_dimensions[1].height = 25
                    worksheet.row_dimensions[2].height = 18
                    worksheet.row_dimensions[4].height = 28
                    
                    for col_num in range(1, 8):
                        cell = worksheet.cell(row=4, column=col_num)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_align
                    
                    total_row_idx = 4 + len(df_excel)
                    
                    for row_idx in range(5, total_row_idx):
                        worksheet.row_dimensions[row_idx].height = 20
                        for col_num in range(1, 8):
                            cell = worksheet.cell(row=row_idx, column=col_num)
                            cell.font = data_font
                            cell.border = data_border
                            
                            if col_num == 1:
                                cell.alignment = center_align
                            elif col_num == 2:
                                cell.alignment = left_align
                            else:
                                cell.alignment = right_align
                                cell.number_format = '#,##0.00'
                                
                    worksheet.row_dimensions[total_row_idx].height = 24
                    for col_num in range(1, 8):
                        cell = worksheet.cell(row=total_row_idx, column=col_num)
                        cell.font = total_font
                        cell.fill = total_fill
                        cell.border = total_border
                        
                        if col_num == 1:
                            cell.alignment = center_align
                        elif col_num == 2:
                            cell.alignment = left_align
                        else:
                            cell.alignment = right_align
                            cell.number_format = '#,##0.00'
                    
                    col_widths = {1: 16, 2: 42, 3: 18, 4: 18, 5: 18, 6: 18, 7: 18}
                    for col_num, width in col_widths.items():
                        col_letter = get_column_letter(col_num)
                        worksheet.column_dimensions[col_letter].width = width
                        
                messagebox.showinfo("Başarılı", "Profesyonel rapor başarıyla Excel'e aktarıldı!")
            except Exception as e:
                messagebox.showerror("Hata", f"Excel kaydedilirken hata oluştu:\n{str(e)}")

    def excele_aktar_ozet(self):
        if self.df is None or self.df.empty:
            messagebox.showwarning("Uyarı", "Dışa aktarılacak veri yok. Önce raporu çalıştırın.")
            return

        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                 filetypes=[("Excel Dosyası", "*.xlsx")],
                                                 title="Genel Toplamı Excel Olarak Kaydet")
        if file_path:
            try:
                start_dt = self.entry_start.get_date().strftime('%d-%m-%Y')
                end_dt = self.entry_end.get_date().strftime('%d-%m-%Y')
                
                toplamlar = pd.DataFrame([{
                    "Açıklama": "GENEL TOPLAM",
                    "Dönem Başı Stok": self.df['Dönem Başı Stok'].sum(),
                    "Dönem İçi Giriş": self.df['Dönem İçi Giriş'].sum(),
                    "Dönem İçi Çıkış": self.df['Dönem İçi Çıkış'].sum(),
                    "Net Hareket": self.df['Net Hareket'].sum(),
                    "Dönem Sonu Stok": self.df['Dönem Sonu Stok'].sum()
                }])
                
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    toplamlar.to_excel(writer, sheet_name='Genel Toplam Özet', index=False, startrow=3)
                    
                    workbook = writer.book
                    worksheet = writer.sheets['Genel Toplam Özet']
                    worksheet.views.sheetView[0].showGridLines = True
                    
                    worksheet['A1'] = "HAFTALIK DEPO DURUM ÖZETİ"
                    worksheet['A1'].font = Font(name='Segoe UI', size=14, bold=True, color='1F497D')
                    
                    worksheet['A2'] = f"Rapor Tarih Aralığı: {start_dt} - {end_dt}"
                    worksheet['A2'].font = Font(name='Segoe UI', size=11, italic=True, color='595959')
                    
                    header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
                    header_fill = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')
                    header_align = Alignment(horizontal='center', vertical='center')
                    
                    for col_num in range(1, 7):
                        cell = worksheet.cell(row=4, column=col_num)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_align
                        
                    total_font = Font(name='Segoe UI', size=11, bold=True)
                    total_fill = PatternFill(start_color='E9EDF4', end_color='E9EDF4', fill_type='solid')
                    total_border = Border(top=Side(border_style="thin", color="1F497D"), bottom=Side(border_style="double", color="1F497D"))
                    
                    worksheet.row_dimensions[4].height = 26
                    worksheet.row_dimensions[5].height = 24
                    
                    for col_num in range(1, 7):
                        cell = worksheet.cell(row=5, column=col_num)
                        cell.font = total_font
                        cell.fill = total_fill
                        cell.border = total_border
                        if col_num == 1:
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                        else:
                            cell.alignment = Alignment(horizontal='right', vertical='center')
                            cell.number_format = '#,##0.00'
                            
                    col_widths = {1: 22, 2: 18, 3: 18, 4: 18, 5: 18, 6: 18}
                    for col_num, width in col_widths.items():
                        col_letter = get_column_letter(col_num)
                        worksheet.column_dimensions[col_letter].width = width
                        
                messagebox.showinfo("Başarılı", "Genel toplam özet dosyası başarıyla kaydedildi!")
            except Exception as e:
                messagebox.showerror("Hata", f"Excel kaydedilirken hata oluştu:\n{str(e)}")

if __name__ == "__main__":
    root = tk.Tk()
    app = StokRaporApp(root)
    root.mainloop()